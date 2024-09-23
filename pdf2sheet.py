import openpyxl
import os
import re
from datetime import datetime

# Define the path and workbook/sheet details
file_path = r"S:\AR Shared Job Files\Job Numbers.xlsx"
sheet_name = "2024"

# Function to find the first available empty row in a specific column
def get_first_empty_row(ws, col):
    row = 1
    while ws[f"{col}{row}"].value:
        row += 1
    return row

# Function to parse the input text and extract data
def parse_input(input_text):
    data = {}
    
    # Strictly following the template you provided
    patterns = {
        "Job Name": r"Job Name:\s*(.+)",
        "Job Address": r"Job Address:\s*(.+)",
        "GC Name": r"GC/Property Owner/Customer Name:\s*(.+)",
        "Architect": r"Architect:\s*(.+)",
        "Job Contract Price": r"Job Contract Price Total:\s*(.+)",
        "Retainage on Contract": r"Retainage on Contract:\s*(.+)",
        "Date Contract Awarded": r"Date Contract Awarded:\s*(.+)"
    }
    
    # Extract data from the input text using regex
    for key, pattern in patterns.items():
        match = re.search(pattern, input_text)
        if match:
            data[key] = match.group(1)
    
    return data

# Function to format the date as "m/dd/yyyy"
def format_date(date_str):
    try:
        # Try to parse and format the date
        date_obj = datetime.strptime(date_str, '%B %d, %Y')
        return date_obj.strftime('%-m/%d/%Y')  # Format as "m/dd/yyyy"
    except ValueError:
        return date_str  # Return as-is if formatting fails

# Function to extract the state abbreviation from the address
def get_state_abbreviation(address):
    # Match a two-letter uppercase state abbreviation (e.g., "NC" for North Carolina)
    match = re.search(r"\b[A-Z]{2}\b", address)
    if match:
        return match.group(0)
    return None

# Function to update the Excel sheet with parsed data
def update_excel(input_text):
    # Load the workbook and select the sheet
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    
    ws = wb[sheet_name]
    
    # Find the first available row in column A
    row = get_first_empty_row(ws, 'D')
    
    # Parse the input text
    parsed_data = parse_input(input_text)
    
    # Update the Excel sheet based on parsed data
    if "Job Name" in parsed_data:
        ws[f"C{row}"] = parsed_data["Job Name"]
    if "Job Address" in parsed_data:
        ws[f"CD{row}"] = parsed_data["Job Address"]
        state_abbreviation = get_state_abbreviation(parsed_data["Job Address"])
        if state_abbreviation:
            ws[f"E{row}"] = state_abbreviation  # Insert state abbreviation in column E
    if "GC Name" in parsed_data:
        ws[f"D{row}"] = parsed_data["GC Name"]
    if "Architect" in parsed_data:
        ws[f"CE{row}"] = parsed_data["Architect"]
    if "Job Contract Price" in parsed_data:
        ws[f"F{row}"] = parsed_data["Job Contract Price"]
    if "Retainage on Contract" in parsed_data:
        ws[f"CF{row}"] = parsed_data["Retainage on Contract"]
    if "Date Contract Awarded" in parsed_data:
        formatted_date = format_date(parsed_data["Date Contract Awarded"])
        ws[f"G{row}"] = formatted_date  # Insert formatted date in column G
    
    # Save the workbook
    wb.save(file_path)
    wb.close()

# Function to accept multiline input
def get_contract_details():
    print("Paste your contract details here (press Enter twice to end):")
    contract_details = []
    while True:
        line = input()
        if line == "":  # End input when the user presses Enter twice
            break
        contract_details.append(line)
    
    return "\n".join(contract_details)

# Use the new function to get the input
input_text = get_contract_details()

# Call the function to update the Excel sheet
update_excel(input_text)
