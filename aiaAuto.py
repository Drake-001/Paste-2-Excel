import openpyxl
import os

# Paths to workbooks and tracking file
jn_path = r"S:\AR Shared Job Files\Job Numbers.xlsx"
at_path = r"S:\01 - Pay Apps 2024\AIA Template.xlsx"
tracking_file = 'last_processed_row.txt'

# Load workbooks
try:
    jn_wb = openpyxl.load_workbook(jn_path)
    at_wb = openpyxl.load_workbook(at_path)
except FileNotFoundError as e:
    print(f"Error: {e}")
    exit(1)

# Select sheets
try:
    jn_sheet = jn_wb["2024"]
    s1_sheet = at_wb["Job 1234"]
    s2_sheet = at_wb["Job SOV"]
except KeyError as e:
    print(f"Error: {e}")
    exit(1)

# Function to find the next row in JN
def find_next_row(sheet, last_row):
    for row in range(last_row + 1, sheet.max_row + 2):
        if sheet[f'A{row}'].value is not None:
            return row
    return None

# Read the last processed row from the tracking file
if os.path.exists(tracking_file):
    with open(tracking_file, 'r') as file:
        last_processed_row = int(file.read().strip())
else:
    last_processed_row = 167  # Start at row 168 if the tracking file doesn't exist

# Get the next row to process in JN
next_row = find_next_row(jn_sheet, last_processed_row)

if next_row is None:
    print("No new rows to process.")
    exit(0)

# Transfer data from JN to AT
s2_sheet[f'B6'] = jn_sheet[f'A{next_row}'].value
s2_sheet[f'D6'] = jn_sheet[f'B{next_row}'].value
s2_sheet[f'E6'] = jn_sheet[f'C{next_row}'].value
s1_sheet[f'B3'] = jn_sheet[f'D{next_row}'].value
s1_sheet[f'E20'] = jn_sheet[f'F{next_row}'].value

# Save the AT workbook with a new name
job_number = s2_sheet[f'D6'].value
job_name = s2_sheet[f'E6'].value
project_manager_initials = s2_sheet[f'B6'].value

# Ensure the values are not None and are strings
if job_number and job_name and project_manager_initials:
    new_filename = f"{job_number} {job_name} {project_manager_initials}.xlsx"
    save_dir = r"S:\01 - Pay Apps 2024\08 Aug"

    # Ensure the directory exists
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    save_path = os.path.join(save_dir, new_filename)
    at_wb.save(save_path)
    print(f"Data transferred from row {next_row} in JN to AT and saved as {new_filename}.")

    # Update the tracking file with the new last processed row
    with open(tracking_file, 'w') as file:
        file.write(str(next_row))
else:
    print("Error: One or more required fields are missing data.")

# Close workbooks
jn_wb.close()
at_wb.close()
